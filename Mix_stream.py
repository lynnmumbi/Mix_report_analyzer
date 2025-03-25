import openpyxl
from openpyxl.styles import  Border, Side
import pandas as pd

from openpyxl.formatting.rule import  IconSetRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from copy import copy
from io import BytesIO

def process_excel(file_path):
    # Loads the workbook
    wb = openpyxl.load_workbook(file_path)


    # Access the first sheet
    sheet = wb['Scoring']

    # Remove columns Q onwards (Columns Q is the 17th column, so retain the first 16 columns)
    if sheet.max_column > 16:
        sheet.delete_cols(17, sheet.max_column - 16)

    # Insert space for headers and descriptions
    sheet.insert_rows(1, amount=22)

    # Format the header section
    def format_header():
        sheet.merge_cells('A2:O2')
        sheet.merge_cells('A3:O3')
        sheet.cell(row=2, column=1).value = "# MONTHLY FLEET REPORT"
        sheet.cell(row=3, column=1).value = "MONTH"

        for row in [2, 3]:
            cell = sheet.cell(row=row, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, underline='single')

        sheet.merge_cells('A5:O5')
        sheet.merge_cells('A6:O6')
        sheet.cell(row=5, column=1).value = "RAG REPORT"
        for row in [5, 6]:
            cell = sheet.cell(row=row, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)

        sheet.merge_cells('A8:I8')
        cell = sheet.cell(row=8, column=1)
        cell.value = "This report shows a classification of drivers in 3 different categories: Green, Amber, and Red."
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        sheet.merge_cells('A9:P9')
        cell = sheet.cell(row=9, column=1)
        cell.value = "We have also made a comparison between the two months, The Red dots are an indication that a vehicle increased on the violations from the previous month, the Light green dots shows an improvement on the different drivers and amber shows no change."
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    from openpyxl.styles import PatternFill, Font, Alignment

    def add_category_description(start_row, name, description, color, sheet):
        # Merge cells for category name
        sheet.merge_cells(f'A{start_row}:D{start_row}')
        cell_name = sheet.cell(row=start_row, column=1)
        cell_name.value = name
        cell_name.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_name.font = Font(bold=True)
        cell_name.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Merge cells for category description
        sheet.merge_cells(f'A{start_row + 1}:I{start_row + 1}')
        cell_desc = sheet.cell(row=start_row + 1, column=1)
        cell_desc.value = description
        cell_desc.alignment = Alignment(wrap_text=True, vertical="top")  # Wrap text enabled


    # Add header and category descriptions for Green, Amber, and Red
    format_header()

    # Green category (10 to 13)
    green_description = (
        "The drivers in this group can serve as mentors or coaches for the rest of the team."

    )
    add_category_description(10, "Green Drivers (0 - 20 violations)", green_description, "339933", sheet)

    # Amber category (15 to 17)
    amber_description = (
        "These are the average drivers who fall in the middle range of performance. While they are neither particularly good "
        "nor bad, we recommend they receive guidance from the top-performing (green) drivers to help them improve."
    )
    add_category_description(15, "Amber Drivers (21 - 40 violations)", amber_description, "FFC000", sheet)

    # Red category (19 to 22)
    red_description = (
        "These drivers require immediate coaching and support. We recommend pairing them with top-performing (green) drivers "
        "for mentorship, along with offering incentives to encourage improvement."
    )
    add_category_description(19, "Red Drivers (above 40 violations)", red_description, "FF0000", sheet)

    # Count and calculate the percentage for each category
    def get_category_counts_and_percentages():
        green_count = 0
        amber_count = 0
        red_count = 0
        total_vehicles = 0

        # Identify the Advanced Score column
        header_row = 23
        advanced_score_col = None

        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=header_row, column=col).value == "Advanced Score":
                advanced_score_col = col
                break

        if advanced_score_col:
            for row in range(header_row + 1, sheet.max_row + 1):
                try:
                    score = sheet.cell(row=row, column=advanced_score_col).value
                    score = int(score)
                    total_vehicles += 1

                    if score <= 20:
                        green_count += 1
                    elif 21 <= score <= 40:
                        amber_count += 1
                    else:
                        red_count += 1
                except (ValueError, TypeError):
                    continue  # Skip invalid scores

        # Calculate percentages
        green_percentage = (green_count / total_vehicles) * 100 if total_vehicles > 0 else 0
        amber_percentage = (amber_count / total_vehicles) * 100 if total_vehicles > 0 else 0
        red_percentage = (red_count / total_vehicles) * 100 if total_vehicles > 0 else 0

        return {
            "Green": {"count": green_count, "percentage": green_percentage},
            "Amber": {"count": amber_count, "percentage": amber_percentage},
            "Red": {"count": red_count, "percentage": red_percentage},
        }

    # Get category counts and percentages
    category_data = get_category_counts_and_percentages()

    # Insert calculated count and percentages
    sheet.merge_cells('A13:H13')
    sheet.cell(row=12, column=1).value = f"This category includes {category_data['Green']['count']} vehicles, accounting for {category_data['Green']['percentage']:.2f}% of the total fleet."
    sheet.merge_cells('A17:H17')
    sheet.cell(row=17, column=1).value = f"This category includes {category_data['Amber']['count']} vehicles, accounting for {category_data['Amber']['percentage']:.2f}% of the total fleet"
    sheet.merge_cells('A21:H21')
    sheet.cell(row=21, column=1).value = f"This category includes {category_data['Red']['count']} vehicles, accounting for {category_data['Red']['percentage']:.2f}% of the total fleet"

    # Add "The table below outlines the vehicles..." description to row 22
    sheet.merge_cells('A22:O22')
    sheet.cell(row=22, column=1).value = "The table below outlines the vehicles in the three categories:"
    sheet.cell(row=22, column=1).alignment = Alignment(horizontal='left', vertical='center')
    sheet.cell(row=22, column=1).font = Font(bold=True)


    #### adding columns to the table ###########33

    # Find the index of the "Advanced Score" column
    header_row = 23
    headers = [cell.value for cell in sheet[header_row]]
    advanced_score_index = headers.index("Advanced Score") + 1  # Convert to 1-based index

    # Insert new columns before and after the "Advanced Score" column
    sheet.insert_cols(advanced_score_index)  # Insert BEFORE "Advanced Score"
    sheet.insert_cols(advanced_score_index + 2)  # Insert AFTER "Advanced Score"

    # Assign headers to the new columns
    sheet.cell(row=header_row, column=advanced_score_index).value = "Previous Month Advanced Score"
    sheet.cell(row=header_row, column=advanced_score_index + 2).value = "Advanced Score Change"


    # Format the table
    # Apply borders to the entire table (from row 23 onwards)
    table_start_row = 23
    for row in sheet.iter_rows(min_row=table_start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )

    # Define the fill color (light blue)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Define the header row number and number of columns
    header_row = 23
    num_columns = sheet.max_column

    # Apply formatting to each header cell
    for col in range(1, num_columns + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill  # Apply light blue background
        cell.alignment = Alignment(wrap_text = True, horizontal= "center", vertical="center")

    # Optionally, auto-adjust column width based on content
    for col in range(1, num_columns + 1):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 20  # Adjust width as needed

    # Apply color coding to the Advanced Score column
    header_row = 23
    advanced_score_col = None

    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=header_row, column=col).value == "Advanced Score":
            advanced_score_col = col
            break

    # Define color fills
    green_fill = PatternFill(start_color="339933", end_color="00FF00", fill_type="solid")  # Green
    amber_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Amber
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

    # Apply color coding to the Advanced Score column
    if advanced_score_col:
        for row in range(header_row + 1, sheet.max_row + 1):
            try:
                score = sheet.cell(row=row, column=advanced_score_col).value
                score = int(score)
                cell = sheet.cell(row=row, column=advanced_score_col)

                if score <= 20:
                    cell.fill = green_fill
                elif 21 <= score <= 40:
                    cell.fill = amber_fill
                else:
                    cell.fill = red_fill
            except (ValueError, TypeError):
                continue  # Skip invalid scores



    ### THE VLOOKUP PART ###
    prev_month_sheet = wb["Previous_month"]  # This contains last month's scores

    # Step 1: Extract "RegistrationNumber" & "Advanced Score" from "Previous_month"
    prev_month_data = {}

    # Find column indexes in "Previous_month"
    prev_headers = [cell.value for cell in prev_month_sheet[1]]
    reg_num_index = prev_headers.index("RegistrationNumber") + 1
    adv_score_index = prev_headers.index("Advanced Score") + 1

    #Store values in a dictionary (like a lookup table)
    for row in prev_month_sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        reg_num = row[reg_num_index - 1]  # Adjust for 0-based index
        adv_score = row[adv_score_index - 1]
        prev_month_data[reg_num] = adv_score  # Store in dict

    # Step 2: Populate "Previous Month Advanced Score" in main table
    # Find column indexes in the main sheet
    main_headers = [cell.value for cell in sheet[23]]  # Header row is 23
    reg_num_col = main_headers.index("RegistrationNumber") + 1
    prev_month_col = main_headers.index("Previous Month Advanced Score") + 1  # Newly added column

    # Loop through rows & fill in previous month's scores
    for row in range(24, sheet.max_row + 1):  # Data starts from row 24
        reg_num = sheet.cell(row=row, column=reg_num_col).value  # Get registration number
        if reg_num in prev_month_data:
            sheet.cell(row=row, column=prev_month_col).value = prev_month_data[reg_num]  # VLOOKUP match


    #############################################################################
    # getting the difference of the two advanced scores
    # Find column indexes
    adv_score_col = main_headers.index("Advanced Score") + 1
    adv_score_change_col = main_headers.index("Advanced Score Change") + 1  # Newly added column

    # Loop through rows & calculate the difference
    for row in range(24, sheet.max_row + 1):  # Data starts from row 24
        prev_score = sheet.cell(row=row, column=prev_month_col).value  # Previous month score
        curr_score = sheet.cell(row=row, column=adv_score_col).value  # Current advanced score

        if prev_score is None or curr_score is None:  # If missing values
            sheet.cell(row=row, column=adv_score_change_col).value = "-"  # Replace with "-"
        else:
            sheet.cell(row=row, column=adv_score_change_col).value = curr_score - prev_score  # Compute difference


    ################# adding icon sets to advanced score change ####################

    # Identify the column letter for "Advanced Score Change"
    advanced_score_change_col = None  # Initialize variable

    for col in sheet.iter_cols(min_row=23, max_row=23):  # Loop through header row
        for cell in col:
            if cell.value == "Advanced Score Change":  # Find the correct column
                advanced_score_change_col = cell.column_letter  # Get letter (e.g., 'E')
                break  # Stop searching once found

    # Ensure column detection works
    if not advanced_score_change_col:
        raise ValueError("Column 'Advanced Score Change' not found!")

    # Convert column values to float to prevent Excel from treating them as text
    for row in sheet.iter_rows(min_row=24, max_row=sheet.max_row,
                               min_col=cell.column, max_col=cell.column):
        for cell in row:
            if isinstance(cell.value, str):  # If stored as text
                try:
                    cell.value = float(cell.value)  # Convert to number
                except ValueError:
                    cell.value = None  # Ensure missing values don't break formatting



    # Define the icon set rule with proper thresholds
    icon_rule = IconSetRule(
        icon_style="3TrafficLights1",  # Uses red, yellow, and green icons
        type="num",
        values=[-100000, 0, 1],
        showValue=True,
        reverse=True
    )

    # Define the range to apply formatting (starting from row 24)
    start_row = 24
    end_row = sheet.max_row
    range_str = f"{advanced_score_change_col}{start_row}:{advanced_score_change_col}{end_row}"

    # Apply conditional formatting
    sheet.conditional_formatting.add(range_str, icon_rule)

    #######################################################################################################################

    # Read the 'Top_N' sheet into a DataFrame
    ws = wb["Top_N"]

    # Define the column indexes to remove (1-based index for openpyxl)
    columns_to_delete = [1, 2]  # A=1, B=2

    # Delete columns from right to left to avoid shifting issues
    for col in sorted(columns_to_delete, reverse=True):
        ws.delete_cols(col)

    # Insert explanation rows in the first 16 rows
    descriptions = [
        ["Top violators:"],
        ["Here are detailed explanations of each violation and potential consequences."],
        [
            "Diagnostic: Fault no Engine RPM: This issue causes the vehicles to miss critical events such as freewheeling and over-revving."],
        [
            "Free wheeling: Freewheeling is likely to cause Gearbox Damage and engine problems in case the driver engages the wrong gear after freewheeling, there’s also increased chances of an accident."],
        [
            "Possible impact: This violation monitors the severity of different impacts and may eventually lead to the shaft of the vehicle breaking, When the drive shaft breaks, the power from the engine can no longer reach the wheels, resulting in a loss of propulsion. Your vehicle may suddenly lose speed and become difficult to accelerate, making it challenging to maintain control."],
        ["Harsh acceleration: This violation reduces tire life and increases fuel consumption."],
        [
            "Harsh braking: This violation causes damage of brake pads & Brake drums, suspension parts and may lead to tire burst and tire reduced tire life."],
        ["Idle - excessive: This results in higher fuel consumption."],
        ["Night Driving: Night driving increases the risk of theft and accidents due to poor visibility."],
        [
            "Over Revving: This violation causes increased tear and wear of the vehicle engine parts and high fuel consumption."],
        ["Over Speeding: This violation results in high fuel consumption, and a high risk of accidents."],
        [
            "3-Axis - Possible Accident (In Trip): This violation monitors the severity of different impacts and may eventually lead to the shaft of the vehicle breaking, When the drive shaft breaks, the power from the engine can no longer reach the wheels, resulting in a loss of propulsion. Your vehicle may suddenly lose speed and become difficult to accelerate, making it challenging to maintain control."],
        ["Over speeding in location: This violation results in high fuel consumption, and a high risk of accidents."],
        [
            "Out of Green Band driving: This violation indicates the engine is operating outside its optimal RPM range, leading to inefficient fuel consumption, increased wear and tear, and higher emissions."],
        [""],
        ["The table below shows the top 3 violators in each RAG label category"]
    ]

    # Insert explanation rows at the beginning
    ws.insert_rows(1, amount=len(descriptions))

    # Apply text to inserted rows
    for i, row_content in enumerate(descriptions, start=1):
        ws.cell(row=i, column=1, value=row_content[0])  # Insert text in column A
        ws.row_dimensions[i].height = 22  # Increase row height for better visibility

    # Apply formatting
    # 1. Merge and center first row across A:H
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 2. Enable text wrapping for description rows (2-16)
    for i in range(2, 17):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=21)
        merged_cell = ws.cell(row=i, column=1)
        merged_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        merged_cell.font = Font(bold=False)  # Ensure it's not bold

    # Identify the new header row after shifting
    header_row_top = len(descriptions) + 1  # Now it's row 17

    # Extract header values from row 17
    header = [cell.value for cell in ws[header_row_top]]

    # color header
    for col in range(1, 5):
        cell = ws.cell(row=header_row_top, column=col)
        cell.fill = header_fill

    # Ensure required headers exist
    required_headers = ["Event", "Metric", "Top 3"]
    missing_headers = [h for h in required_headers if h not in header]

    if missing_headers:
        print(f"❌ Error: Missing required columns: {missing_headers} in row {header_row}")
        print(f"Extracted Headers: {header}")
        wb.close()
        raise ValueError("Required headers not found in expected row.")

    # Find column indexes dynamically
    event_col = header.index("Event") + 1
    metric_col = header.index("Metric") + 1
    top3_col = header.index("Top 3") + 1

    # Recalculate last_row based on the last non-empty row after row deletions
    last_row = header_row_top  # Start from header row
    for row in range(header_row_top, ws.max_row + 1):
        if any(ws.cell(row=row, column=col).value for col in range(1, 5)):
            last_row = row

    if last_row < header_row_top:
        raise ValueError("No valid data rows found after deletions.")

    # Define event-metric mapping
    event_metric_mapping = {
        "Night Driving": "DistanceKilometers",
        'Night Driving 11pm- 4am': "DistanceKilometers",
        "PTRM - Night driving | Alarm": "DistanceKilometers",
        'Night driving': "DistanceKilometers",
        'Night Driving Time': "DistanceKilometers",
        'Ragos Night Driving': "DistanceKilometers",
        'Menengai Night Driving': "DistanceKilometers",
        "3-Axis - Possible Accident (In Trip)": "TotalOccurances",
        'FreeWheeling': "DistanceKilometers",
        "Free  Wheeling": "DistanceKilometers",
        "Free wheel violations": "DistanceKilometers",
        'Free Wheeling Violation': "DistanceKilometers",
        "Free Wheeling": "DistanceKilometers",
        'Free wheeling in Neutral': "DistanceKilometers",
        'Free wheeling': "DistanceKilometers",
        'Free Wheeling at Neutral': "DistanceKilometers",
        'Freewheeling': 'DistanceKilometers',
        'Roy Trans Free Wheeling': 'DistanceKilometers',
        'Gazlin Freewheeling': "DistanceKilometers",
        "Harsh braking": "TotalOccurances",
        "PTRM - Harsh braking 14km/h/s | Alarm": "TotalOccurances",
        'PTRM - Harsh braking 10km/h/s | Record (#/100 km)': "TotalOccurances",
        "Over Revving": "Duration",
        "Over revving": "Duration",
        "Idle - excessive": "Duration",
        'Idle - excessive Timsales': "Duration",
        'Idle - excessive above 10 Mins': "Duration",
        "Harsh acceleration": "TotalOccurances",
        "PTRM - Harsh acceleration 10km/h/s | Alarm": "TotalOccurances",
        "Over Speeding": "DistanceKilometers",
        "Over speeding": "DistanceKilometers",
        'Over speeding above 80km/h': "DistanceKilometers",
        'Over speeding above 80 km/h': "DistanceKilometers",
        "PTRM - Over speeding > 84km/h | Alarm": "DistanceKilometers",
        "PTRM - Overspeeding in 30kph Speed Zone | Alarm": "DistanceKilometers",
        "PTRM - Overspeeding in 50kph Speed Zone | Alarm": "DistanceKilometers",
        'Over speeding above 70km/h': "DistanceKilometers",
        'Over speeding > 74km/h | Record': "DistanceKilometers",
        'Road Speed Overspeeding': "DistanceKilometers",
        'Ragos - Over Speeding 70km/h': "DistanceKilometers",
        'PTRM - Over speeding | Alarm': "DistanceKilometers",
        "Diagnostic: Fault no Engine RPM": "TotalOccurances",
        "Over speeding in location": "DistanceKilometers",
        "Possible impact": "TotalOccurances",
        "Out of Green Band driving": "TotalOccurances",
        'PTRM - Harsh acceleration 6km/h/s | Record': "TotalOccurances",
        'PTRM - Harsh braking 10km/h/s | Record': "TotalOccurances",
        'PTRM - Night driving | Record': "DistanceKilometers",
        'RAIMDF Night Driving': 'DistanceKilometers',
        'No Go Zone Alert': "TotalOccurances",
        'Rubis Idle Excessive': "Duration",
        'EXPEDITERS NIGHT DRIVING 2100Hrs  - 0500Hrs': "DistanceKilometers",
        'PTRM - Harsh braking | Alarm': 'TotalOccurances',
        'PTRM - Harsh acceleration | Alarm': 'TotalOccurances',
        '3-Axis- Possible Accident Impact': 'TotalOccurances',
        'Idle - excessive above 15 Mins': 'Duration'

    }

    # Filter rows based on event-metric mapping
    rows_to_delete = []
    for row in range(header_row_top + 1, last_row + 1):  # Start from data rows
        event = ws.cell(row=row, column=event_col).value
        metric = ws.cell(row=row, column=metric_col).value

        # Ensure values are valid before comparison
        if event is None or metric is None:
            continue  # Skip empty rows

        if event not in event_metric_mapping or metric != event_metric_mapping[event]:
            rows_to_delete.append(row)

    # Delete rows from bottom to top to avoid shifting issues
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

    # Recalculate last_row based on the last non-empty row after row deletions
    last_row = header_row_top  # Start from header row

    for row in range(header_row_top, ws.max_row + 1):
        if any(ws.cell(row=row, column=col).value for col in
               range(1, 5)):  # Check if any of the first 4 columns have data
            last_row = row  # Update last_row to the last row that contains data

    # Update remaining rows: Fill blank "Top 3" cells
    for row in range(header_row_top + 1, ws.max_row + 1):
        top3_cell = ws.cell(row=row, column=top3_col)
        if not top3_cell.value:
            top3_cell.value = "There were no incidences recorded in this category"

    # Apply borders to the table
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply borders to each cell in the table
    for row in range(header_row_top, last_row + 1):
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border

    # Save the modified workbook
    # wb.save(file_path)
    wb.close()

    print("✅ 'Top_N' sheet successfully updated.")
    ################################ANALYSIS#######################################
    # lets do the analysis
    analysis_sheet = wb["Analysis"]

    #read headers
    headers_analysis = [cell.value for cell in analysis_sheet[1]]
    print("Headers have been found!", headers_analysis)

    # Read headers dynamically
    headers_analysis = [cell.value for cell in analysis_sheet[1]]

    # identifying columns dynamically
    fixed_columns = ["AssetDescription", "RegistrationNumber", "Distance (KM)"]
    percentage_columns =["Over Speeding Distance (%)",
                         "Over speeding Distance (%)",
                         "Over speeding above 80km/h Distance (%)",
                         'Free wheel violations Distance (%)',
                         "PTRM - Overspeeding in 30kph Speed Zone | Alarm Distance (%)",
                         "PTRM - Overspeeding in 50kph Speed Zone | Alarm Distance (%)",
                         "PTRM - Over speeding > 84km/h | Alarm Distance (%)",
                         'Over speeding > 74km/h | Record Distance (%)',
                         "Free  Wheeling Distance (%)",
                         "Free wheel violations Distance (%)",
                         'FreeWheeling Distance (%)',
                         'Free wheeling Distance (%)',
                         'Free Wheeling at Neutral Distance (%)',
                         'Roy Trans Free Wheeling Distance (%)',
                         "Harsh braking (#/100 km)",
                         "PTRM - Harsh braking 14km/h/s | Alarm (#/100 km)",
                         "Over speeding in location Distance (%)",
                         "Free Wheeling Distance (%)",
                         'Free Wheeling Violation Distance (%)',
                         'Free wheel violations Distance (%)'
                         'Free wheeling in Neutral Distance (%)',
                         'Freewheeling Distance (%)',
                         'Free wheel violations Distance (%)',
                         'PTRM - Harsh braking 10km/h/s | Record (#/100 km)',
                         'PTRM - Harsh braking | Alarm (#/100 km)',
                         'Over speeding above 80 km/h Distance (%)',
                         'Gazlin Freewheeling Distance (%)',
                         'Over speeding above 70km/h Distance (%)',
                         'Road Speed Overspeeding Distance (%)',
                         'Ragos - Over Speeding 70km/h Distance (%)',
                         'PTRM - Over speeding | Alarm Distance (%)',
                         'Over speeding above 80 km/h Distance (%)',
                                         ]


    # Extract column indices
    col_indices = {col: headers_analysis.index(col) for col in fixed_columns + percentage_columns if col in headers_analysis}

    # Define formatting styles
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Process each percentage violation column separately
    for violation_col in percentage_columns:
        if violation_col in col_indices:
            # Ensure sheet name does not exceed 31 characters
            sheet_name = violation_col.replace(" (%)", "").replace("/", "-")
            sheet_name = sheet_name[:31]  # Truncate to max allowed length

            # Create a new sheet
            new_sheet = wb.create_sheet(sheet_name)

            # Define the selected headers for this sheet
            selected_headers = fixed_columns + [violation_col]

            # Apply formatting to header row
            for col_idx, header in enumerate(selected_headers, start=1):
                cell = new_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            # Extract relevant data
            for row in analysis_sheet.iter_rows(min_row=2, values_only=True):
                new_row = [row[col_indices[col]] for col in selected_headers]
                new_sheet.append(new_row)

            # Auto-adjust column widths
            for col_idx, col_name in enumerate(selected_headers, start=1):
                new_sheet.column_dimensions[new_sheet.cell(row=1, column=col_idx).column_letter].width = max(15, len(col_name) + 2)

    #placing the tables into new sheet
    # Create a new sheet for all tables
    combined_sheet = wb.create_sheet("Violation Summary")

    # Column position tracker
    start_col = 1  # Start at column A
    start_row = 4  # Set the starting row for headers (row 4)

    # Merge and format row 1 - VIOLATION SUMMARY
    title_cell = combined_sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=8)
    title_cell = combined_sheet.cell(row=1, column=4, value="VIOLATION SUMMARY")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Increase row height for row 1
    combined_sheet.row_dimensions[1].height = 30

    # Merge and format row 2 - The following tables show violations...
    description_cell = combined_sheet.merge_cells(start_row=2, start_column=4, end_row=2, end_column=8)
    description_cell = combined_sheet.cell(row=2, column=4, value="The following tables show violations that were captured.")
    description_cell.font = Font(size=11)
    description_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Increase row height for row 2
    combined_sheet.row_dimensions[2].height = 25

    for violation_col in percentage_columns:
        if violation_col in col_indices:
            # Define the headers for this section
            selected_headers = fixed_columns + [violation_col]

            # Filter rows: Keep only if at least one value in violation_col is greater than 0
            filtered_rows = [
                [row[col_indices[col]] for col in selected_headers]
                for row in analysis_sheet.iter_rows(min_row=2, values_only=True)
                if row[col_indices[violation_col]] > 0  # Only keep if violation > 0
            ]

            if not filtered_rows:  # Skip if all values are 0 or less
                continue

            # Sort rows by violation(%) in ascending order
            filtered_rows.sort(key=lambda x: x[-1])  # Sorting by the last column (violation %)

            # Format violation(%) to 3 decimal places
            for row in filtered_rows:
                row[-1] = round(row[-1], 3)

            # Write headers in the correct position (row 4 now)
            for col_idx, header in enumerate(selected_headers):
                cell = combined_sheet.cell(row=start_row, column=start_col + col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            # Write data under the headers (starting row 5)
            row_num = start_row + 1  # Start data immediately after the headers (row 5)
            for new_row in filtered_rows:
                for col_idx, value in enumerate(new_row):
                    cell = combined_sheet.cell(row=row_num, column=start_col + col_idx, value=value)
                    cell.border = border  # Apply borders
                row_num += 1  # Move to next row

            # Adjust column widths
            for col_idx, col_name in enumerate(selected_headers):
                col_letter = combined_sheet.cell(row=start_row, column=start_col + col_idx).column_letter
                combined_sheet.column_dimensions[col_letter].width = max(15, len(col_name) + 2)

                # Apply conditional formatting to violation(%) column
                violation_col_idx = start_col + len(fixed_columns)  # Position of the violation column
                color_scale_rule = ColorScaleRule(
                    start_type="min", start_color="339933",  # Green for low values
                    mid_type="percentile", mid_value=50, mid_color="FFFF00",  # Yellow for medium values
                    end_type="max", end_color="FF0000"  # Red for high values
                )
                combined_sheet.conditional_formatting.add(
                    f"{combined_sheet.cell(row=start_row + 1, column=violation_col_idx).coordinate}:"  # Data starts at row 5
                    f"{combined_sheet.cell(row=row_num - 1, column=violation_col_idx).coordinate}",
                    color_scale_rule
                )

            # Move start column to next table position (+1 for the empty column)
            start_col += len(selected_headers) + 1




    ############################ utilization #######################################
    # Load workbook
    utilization_sheet = wb["Utilization"]

    # Shift the table down so that it starts from row 13
    # Identify the first row with actual data (assumed directly after headers)
    data_start_row = 1  # Adjust if needed
    data_end_row = utilization_sheet.max_row  # Get last used row

    # Calculate how many rows to shift (new start is row 13)
    shift_by = 13 - data_start_row


    if shift_by > 0:
        for row in range(data_end_row, data_start_row - 1, -1):  # Move from bottom to top
            for col in range(1, utilization_sheet.max_column + 1):
                old_cell = utilization_sheet.cell(row=row, column=col)
                new_cell = utilization_sheet.cell(row=row + shift_by, column=col)

                # Copy values and formatting
                new_cell.value = old_cell.value
                new_cell.font = copy(old_cell.font)
                new_cell.fill = copy(old_cell.fill)
                new_cell.alignment = copy(old_cell.alignment)
                new_cell.border = copy(old_cell.border)

                # Clear old cell
                old_cell.value = None

    print(f"Table successfully shifted down. Data now starts at row 13.")



    # Clear existing formatting in the first 12 rows
    for row in range(1, 13):
        for cell in utilization_sheet[row]:
            cell.font = Font(bold=False)
            cell.alignment = Alignment(horizontal="left")

    # Remove borders from column A, rows 1 to 10
    for row in range(1, 11):
        utilization_sheet[f"A{row}"].border = Border()


    # Merge row 1 from A to AJ
    merge_range = f"A1:AJ1"
    utilization_sheet.merge_cells(merge_range)

    # Remove borders from the first row
    for col in range(1, utilization_sheet.max_column + 1):
        utilization_sheet.cell(row=1, column=col).border = Border()


    # Format merged cell
    cell = utilization_sheet["A1"]
    cell.value = "DAILY UTILIZATION REPORT"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True, underline="single")

    #increasing row height
    utilization_sheet.row_dimensions[1].height = 25

    #leave out row 2 and 3
    for row in range(2,4):
        for cell in utilization_sheet[row]:
            cell.value= None

    # Define color fills
    red_fill_util = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    amber_fill_util = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
    yellow_fill_util = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill_util= PatternFill(start_color="008000", end_color="008000", fill_type="solid")  # Dark Green

    # Define border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Define descriptions
    descriptions = [
        "Less Than 0.1 km",
        "Less Than 10 km",
        "Less Than 100 Km",
        "More Than 100 Km"
    ]



    #color list
    color_fills = [red_fill_util, amber_fill_util, yellow_fill_util,green_fill_util]

    # Find "Weekday Distance (km)" column dynamically
    table_start_row_uti_col = 14  # Assuming headers start at row 13
    header_row_uti_number = 13
    weekday_distance_col_index = None

    for cell in utilization_sheet[header_row_uti_number]:
        if cell.value == "Weekday Distance (km)":
            weekday_distance_col_index = cell.column
            break

    if weekday_distance_col_index is None:
        raise ValueError("Could not find 'Weekday Distance (km)' column.")

    # Set the last column to color (just before "Weekday Distance (km)")
    table_start_col = 1
    table_end_col = weekday_distance_col_index - 1

    # Rows to color - start from row 14 (after headers), or your desired start row
    table_start_row = 14
    table_end_row = utilization_sheet.max_row - 1  # Exclude last row if it's totals


    # Apply color coding based on logic you want (adjust as needed)

    for row in utilization_sheet.iter_rows(min_row=table_start_row, max_row=table_end_row,
                                           min_col=table_start_col, max_col=table_end_col):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value >= 100:
                    cell.fill = green_fill_util
                elif cell.value >= 10:
                    cell.fill = yellow_fill_util
                elif cell.value >= 0.1:
                    cell.fill = amber_fill_util
                else:
                    cell.fill = red_fill_util









    # Apply colors, descriptions, and borders
    for i, desc in enumerate(descriptions):
        row = i + 4  # Row 4 to 7
        cell_q = utilization_sheet[f"Q{row}"]
        cell_q.fill = color_fills[i]  # Apply color
        cell_q.border = thin_border  # Apply border

        utilization_sheet.merge_cells(f"R{row}:V{row}")  # Merge R to V
        cell_rv = utilization_sheet[f"R{row}"]
        cell_rv.value = desc  # Add description
        cell_rv.alignment = Alignment(horizontal="left")  # Align left
        cell_rv.border = thin_border  # Apply border to merged cells

        # Debugging statements to check values
        print(f"Setting description at row {row}: {desc}")
        print(f"Cell before merge: {utilization_sheet[f'R{row}'].value}")
        print(f"Cell after merge: {utilization_sheet[f'R{row}'].value}")

        # Apply borders to merged range manually (since only the top-left cell gets a border by default)
        for col in range(ord("R"), ord("V") + 1):  # Iterate from column R to V
            utilization_sheet[f"{chr(col)}{row}"].border = thin_border



    # Merge cells from Q9 to AA11
    utilization_sheet.merge_cells("Q9:AA9")
    utilization_sheet.merge_cells("Q10:AA10")
    utilization_sheet.merge_cells("Q11:AA11")



    # Identify headers dynamically from row 13
    headers_utilization = [cell.value for cell in utilization_sheet[13] if cell.value]  # Remove None values

    # Get column count
    total_columns = len(headers_utilization)

    # Identify key columns
    vehicle_col = 1  # First column is vehicle column
    total_distance_col = total_columns - 2  # Third last column (always)

    # Find the first non-empty row (Headers must be on or after row 13)
    header_row_util = 13  # Start checking from row 13
    while header_row_util <= utilization_sheet.max_row:
        headers_utilization = [cell.value for cell in utilization_sheet[header_row_util] if cell.value]  # Remove empty values
        if headers_utilization:  # Found non-empty row
            break
        header_row_util += 1
    else:
        raise ValueError("No headers found in the sheet.")


    # Identify first and last data rows
    table_start_row = 14  # Data starts after header
    table_end_row = utilization_sheet.max_row  # Last row with data

    # Convert to Excel column letter
    total_distance_col_letter = utilization_sheet.cell(row=13, column=total_distance_col).column_letter

    # Get all distances except the totals row
    fleet_distances = [
        (utilization_sheet[f"A{row}"].value, utilization_sheet[f"{total_distance_col_letter}{row}"].value or 0)
        for row in range(table_start_row, table_end_row)  # Excluding totals
    ]

    # Sort by distance to find least and most utilized vehicles
    fleet_distances_sorted = sorted(fleet_distances, key=lambda x: x[1])  # Sort by distance


    least_utilized_vehicle, least_distance = fleet_distances_sorted[0]
    most_utilized_vehicle, most_distance = fleet_distances_sorted[-1]

    # Compute fleet average, excluding totals row
    fleet_average = sum([dist for _, dist in fleet_distances]) / len(fleet_distances) if fleet_distances else 0

    # Merge cells from Q9 to AA11
    utilization_sheet.merge_cells("Q9:AA9")
    utilization_sheet.merge_cells("Q10:AA10")
    utilization_sheet.merge_cells("Q11:AA11")

    # Set values
    utilization_sheet["Q9"].value = f"The least utilized vehicle was {least_utilized_vehicle} with {least_distance} KM"
    utilization_sheet["Q10"].value = f"The most utilized vehicle was {most_utilized_vehicle} with {most_distance} KM"
    utilization_sheet["Q11"].value = f"The average distance covered by each vehicle in the fleet was {fleet_average:.1f} KM"



    # Format text: bold for key terms
    for row in range(9, 12):
        cell = utilization_sheet[f"Q{row}"]
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.font = Font(bold=True)

    # color coding the uti table
    #damn this alot of code
    #already defined the border style
    table_start_col = 1
    table_end_col_with_total_dis=utilization_sheet.max_column
    table_end_col = utilization_sheet.max_column - 5  # Exclude last 5 columns

    # Apply borders to all data cells
    for row in range(table_start_row, table_end_row + 1):
        for col in range(table_start_col, table_end_col_with_total_dis + 1):
            utilization_sheet.cell(row=row, column=col).border = thin_border

    # Wrap text in headers & color headers light blue
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")


    for col in range(table_start_col, table_end_col_with_total_dis + 1):
        header_cell = utilization_sheet.cell(row=header_row_util, column=col)
        header_cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        header_cell.fill = header_fill
        header_cell.border = thin_border  # Ensure headers also have borders

    # Define color mapping
    color_mapping = {
        "Red": "FF0000",      # Less than 0.1 km
        "Amber": "FFA500",    # Less than 10 km
        "Yellow": "FFFF00",   # Less than 100 km
        "Green": "008000"     # More than 100 km
    }

    # Get the range for data (excluding totals and last 5 columns)
    for row in range(table_start_row, table_end_row ):  # Loop through data rows
        for col in range(table_start_col, table_end_col - 4):  # Exclude last 5 columns
            cell = utilization_sheet.cell(row=row, column=col)

            # Convert cell value to a number (if possible)
            try:
                cell_value = float(cell.value)  # Convert to float if possible
            except (TypeError, ValueError):
                continue  # Skip non-numeric or empty cells

            # Apply color coding based on conditions
            if cell_value < 0.1:
                cell.fill = PatternFill(start_color=color_mapping["Red"], fill_type="solid")
            elif cell_value < 10:
                cell.fill = PatternFill(start_color=color_mapping["Amber"], fill_type="solid")
            elif cell_value < 100:
                cell.fill = PatternFill(start_color=color_mapping["Yellow"], fill_type="solid")
            else:
                cell.fill = PatternFill(start_color=color_mapping["Green"], fill_type="solid")





    #Onto Fuel Analysis

    # read sheet
    fuel_sheet = wb["Fuel"]

    # Shift the table down
    # Identify the first row with actual data (assumed directly after headers)
    fuel_start_row = 1  # Adjust if needed
    fuel_end_row = fuel_sheet.max_row  # Get last used row

    # Calculate how many rows to shift
    shift_by = 5 - fuel_start_row


    if shift_by > 0:
        for row in range(fuel_end_row, fuel_start_row - 1, -1):  # Move from bottom to top
            for col in range(1, fuel_sheet.max_column + 1):
                fuel_old_cell = fuel_sheet.cell(row=row, column=col)
                fuel_new_cell = fuel_sheet.cell(row=row + shift_by, column=col)

                # Copy values and formatting
                fuel_new_cell.value = fuel_old_cell.value
                fuel_new_cell.font = copy(fuel_old_cell.font)
                fuel_new_cell.fill = copy(fuel_old_cell.fill)
                fuel_new_cell.alignment = copy(fuel_old_cell.alignment)
                fuel_new_cell.border = copy(fuel_old_cell.border)

                # Clear old cell
                fuel_old_cell.value = None

    print(f"fuel Table successfully shifted down. Data now starts at row 4.")

    # Clear existing formatting in the first 3 rows
    for row in range(1, 4):
        for cell in fuel_sheet[row]:
            cell.font = Font(bold=False)
            cell.alignment = Alignment(horizontal="left")

    # Remove borders from column A, rows 1 to 3
    for row in range(1, 4):
        fuel_sheet[f"A{row}"].border = Border()


    # Merge row 1 from A to D
    merge_range = f"A1:I1"
    fuel_sheet.merge_cells(merge_range)

    # Remove borders from the first row
    for col in range(1, fuel_sheet.max_column + 1):
        fuel_sheet.cell(row=1, column=col).border = Border()

    # Merge row 3
    merge_range = f"A3:I3"
    fuel_sheet.merge_cells(merge_range)

    # Format merged cell
    cell = fuel_sheet["A1"]
    cell.value = "FUEL REPORT"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True, underline="single")

    #increasing row height
    fuel_sheet.row_dimensions[1].height = 25

    # Leave row 2 blank
    for cell in fuel_sheet[2]:
        cell.value = None

    # Format merged cell
    cell = fuel_sheet["A3"]
    cell.value = ("The table below displays the fuel consumption for vehicles equipped with CAN BUS. "
                  "Vehicles highlighted in yellow may be experiencing CAN issues and should be inspected. "
                  "Additionally, we offer comprehensive fuel management solutions—please contact us to arrange a demo "
                  "and explore how these solutions can benefit your fleet.")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


    #increasing row height
    fuel_sheet.row_dimensions[3].height = 55

    # Leave row 4 blank
    for cell in fuel_sheet[4]:
        cell.value = None

    # Remove borders FROM ROW 4
    for row in range(4, 5):
        fuel_sheet[f"A{row}"].border = Border()

    # Identify the header row (Assuming 5th row is headers)
    header_row_fuel = 5

    # Identify the last column dynamically
    last_col_idx = fuel_sheet.max_column

    #get fuel consumption level
    headers_fuel = [cell.value for cell in fuel_sheet[header_row]]

    # Apply light blue header fill
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for col in range(1,5):
        cell = fuel_sheet.cell(row=header_row_fuel, column = col)
        cell.fill = header_fill


    # Apply borders to entire table
    thin_border = Border(left=Side(style="thin"),
                         right=Side(style="thin"),
                         top=Side(style="thin"),
                         bottom=Side(style="thin"))



    #Deleting weird consumption rates( the inf)  and highlighting them
    yellow_fill_fuel = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    rows_to_delete = []
    data_start_row = 6  # Assuming your data starts from row 5

    for row in range(data_start_row, fuel_sheet.max_row + 1):
        cell_value = str(fuel_sheet.cell(row=row, column=4).value).strip()  # Read column D (4th column)

        if cell_value in ["N/A", "inf", "", "None"]:
            rows_to_delete.append(row)

    # Delete rows in reverse order (so row indexes don’t shift)
    for row in reversed(rows_to_delete):
        fuel_sheet.delete_rows(row)

    print(f"Deleted {len(rows_to_delete)} rows where column D contained N/A, inf, blank, or null.")


    #highlighting
    # Define the range (only columns 1 to 4, rows starting from data_start_row)
    for row in range(data_start_row, fuel_sheet.max_row + 1):
        cell_value = fuel_sheet.cell(row=row, column=4).value

        # Ensure the cell contains something and is numeric
        if cell_value is not None and isinstance(cell_value, (int, float)):
            if float(cell_value) < 1 or float(cell_value) > 5:
                for col in range(1, 5):  # Columns A to D (1 to 4)
                    fuel_sheet.cell(row=row, column=col).fill = yellow_fill_fuel

    # Loop through data rows and format columns 2, 3, 4 to 2 decimal places
    for row in range(data_start_row, fuel_sheet.max_row + 1):
        for col in range(2, 5):  # Columns B, C, D (Distance, Fuel Used, Consumption Rate)
            cell = fuel_sheet.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):  # Only format numbers
                cell.value = float(f"{cell.value:.2f}")


    # Apply borders only within data range (from header_row_fuel to last_data_row, and only for columns 1-5)
    for row in fuel_sheet.iter_rows(
        min_row=header_row_fuel,
        max_row=fuel_end_row,  # Ensures it stops at the last data row
        min_col=1,
        max_col=5  # Restricting to first 5 columns
    ):
        for cell in row:
            if cell.value:  # Only apply borders to non-empty cells
                cell.border = thin_border





    #############################sheets to remain###############################################
    # List of sheets to retain in order
    sheets_to_keep_and_order = ["Scoring", "Filtered_Top3", "Violation Summary", "Utilization", "Fuel","Top_N"]


    # Iterate over a list to avoid modifying the workbook object while iterating
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in sheets_to_keep_and_order:
            wb.remove(wb[sheet_name])

    #Reorder the remaining sheets
    for index, sheet_name in enumerate(sheets_to_keep_and_order):
        if sheet_name in wb.sheetnames:
            wb.move_sheet(sheet_name, offset=index - wb.sheetnames.index(sheet_name))

    print("✅ Go see what you did baby girl!")
    # Save to a BytesIO object instead of a file
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream
